from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
from database.connection import DatabaseConnection
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                            QPushButton, QLineEdit, QComboBox, QDateEdit,
                            QGroupBox, QFormLayout, QTableWidget, QTableWidgetItem, QFileDialog, QCalendarWidget, QSpacerItem, QSizePolicy)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtWidgets import QApplication


class ReporteExcelGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.setup_styles()
        
    def setup_styles(self):
        self.corporate_colors = {
            'primary': '1F4E78',
            'secondary': '2F75B5',
            'accent': 'BDD7EE',
            'header': '305496',
            'subheader': '8EA9DB'
        }
        
        self.title_style = NamedStyle(name='title_style')
        self.title_style.font = Font(name='Arial', size=24, bold=True, color='FFFFFF')
        self.title_style.fill = PatternFill(start_color=self.corporate_colors['primary'], 
                                            end_color=self.corporate_colors['primary'],
                                            fill_type='solid')
        self.title_style.alignment = Alignment(horizontal='center', vertical='center')
        self.title_style.border = Border(bottom=Side(style='medium', color=self.corporate_colors['secondary']))
        
        self.header_style = NamedStyle(name='header_style')
        self.header_style.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill(start_color=self.corporate_colors['header'],
                                            end_color=self.corporate_colors['header'],
                                            fill_type='solid')
        self.header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.header_style.border = Border(left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'), 
                                        bottom=Side(style='thin'))
        
        self.data_style = NamedStyle(name='data_style')
        self.data_style.font = Font(name='Arial', size=11)
        self.data_style.alignment = Alignment(horizontal='center', vertical='center')
        self.data_style.border = Border(left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'), 
                                        bottom=Side(style='thin'))

        self.currency_style = NamedStyle(name='currency_style')
        self.currency_style.font = Font(name='Arial', size=11)
        self.currency_style.number_format = '"$"#,##0.00'
        self.currency_style.alignment = Alignment(horizontal='center', vertical='center')
        self.currency_style.border = Border(left=Side(style='thin'), 
                                            right=Side(style='thin'), 
                                            top=Side(style='thin'), 
                                            bottom=Side(style='thin'))

    def generate_report(self, data_inventario, data_entradas=None, data_salidas=None, start_date=None, end_date=None, report_type="Mensual"):
        # Asegurarse de que los datos sean listas
        data_inventario = list(data_inventario) if data_inventario else []
        data_entradas = list(data_entradas) if data_entradas else []
        data_salidas = list(data_salidas) if data_salidas else []
        
        self.ws.title = "Inventario"
        start_row = self.add_logo()
        
        self.ws.merge_cells(f'A{start_row}:J{start_row}')
        title_cell = self.ws[f'A{start_row}']
        title_cell.value = "MULTIMUEBLES LA PLATA"
        title_cell.style = self.title_style
        
        # Subtítulo con período
        period_row = start_row + 1
        self.ws.merge_cells(f'A{period_row}:J{period_row}')
        period_cell = self.ws[f'A{period_row}']
        period_cell.value = self._get_report_period(report_type, start_date, end_date)
        period_cell.style = self.header_style
        
        # Encabezados de la tabla
        headers = ["Código", "Descripción", "Entradas", "Salidas", "Stock", 
                    "Precio Compra", "Precio Venta", "Valor Compra Total", 
                    "Valor Venta Total", "Utilidad"]
        
        header_row = period_row + 2
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=header_row, column=col)
            cell.value = header
            cell.style = self.header_style
            self.ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Datos de inventario
        if data_inventario:
            for row_idx, row_data in enumerate(data_inventario, header_row + 1):
                row_data = list(row_data)  # Convertir la fila a lista si no lo es
                for col_idx, value in enumerate(row_data, 1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx in [6, 7, 8, 9, 10]:  # Aplicar estilo de moneda a columnas específicas
                        cell.style = self.currency_style
                    else:
                        cell.style = self.data_style
        
        # Crear hojas adicionales si hay datos
        if data_entradas:
            self._create_movement_sheet("Entradas", data_entradas)
        
        if data_salidas:
            self._create_movement_sheet("Salidas", data_salidas)
        
        # Ajustar anchos de columna y aplicar bordes
        self.autoajustar_columnas(self.ws)
        self.aplicar_bordes(self.ws, self.ws.max_row, self.ws.max_column)
        
        return self.wb

    def _create_movement_sheet(self, sheet_name, data):
        ws = self.wb.create_sheet(title=sheet_name)
        start_row = self.add_logo(ws)
        
        # Título principal
        ws.merge_cells(f'A{start_row}:F{start_row}')
        ws[f'A{start_row}'] = "MULTIMUEBLES LA PLATA"
        ws[f'A{start_row}'].style = self.title_style
        
        # Subtítulo específico
        period_row = start_row + 1
        ws.merge_cells(f'A{period_row}:F{period_row}')
        ws[f'A{period_row}'] = sheet_name.upper()  # Ej: "ENTRADAS"
        ws[f'A{period_row}'].style = self.header_style
        
        # Encabezados de la tabla
        headers = ["Fecha", "Código", "Descripción", "Cantidad", "Precio", "Valor Total"]
        
        header_row = period_row + 2
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.style = self.header_style
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Datos de movimientos
        for row_idx, row_data in enumerate(data, header_row + 1):
            row_data = list(row_data)  # Convertir la fila a lista si no lo es
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                if col_idx in [5, 6]:  # Aplicar estilo de moneda a columnas específicas
                    cell.style = self.currency_style
                else:
                    cell.style = self.data_style
        
        # Ajustar anchos de columna y aplicar bordes
        self.autoajustar_columnas(ws)
        self.aplicar_bordes(ws, ws.max_row, ws.max_column)

    def add_logo(self, worksheet=None):
        logo_path = 'src/resources/img/logo-multimuebles.jpg'
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 150
            img.height = 75
            if worksheet:
                worksheet.add_image(img, 'A1')
            else:
                self.ws.add_image(img, 'A1')
            return 4
        return 1

    def _get_report_period(self, report_type, start_date, end_date):
        if not start_date:
            end_date = datetime.now()
            if report_type == "Mensual":
                start_date = end_date.replace(day=1)
            elif report_type == "Trimestral":
                start_date = end_date - timedelta(days=90)
            elif report_type == "Anual":
                start_date = end_date.replace(month=1, day=1)
            else:
                start_date = end_date - timedelta(days=30)
        
        days = (end_date - start_date).days
        months = days // 30
        return f"Periodo: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')} ({days} días, {months} meses)"

    def autoajustar_columnas(self, worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def aplicar_bordes(self, worksheet, max_row, max_col):
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = self.data_style.border

def fetch_inventory_data(db_connection):
    cursor = db_connection.cursor()
    cursor.execute("""
        SELECT codigo, descripcion, entradas_totales, salidas_totales, 
                stock, precio_compra, precio_venta 
        FROM productos
    """)
    data = []
    for row in cursor.fetchall():
        codigo, descripcion, entradas, salidas, stock, precio_compra, precio_venta = row
        precio_compra_total = entradas * precio_compra
        precio_venta_total = salidas * precio_venta
        valor_total = precio_compra_total + precio_venta_total
        data.append([
            codigo, descripcion, entradas, salidas, stock, 
            precio_compra, precio_venta, precio_compra_total, 
            precio_venta_total, valor_total
        ])
    return data

def generate_inventory_report(db_connection, report_type="Mensual"):
    data = fetch_inventory_data(db_connection)
    generator = ReporteExcelGenerator()
    wb = generator.generate_report(data, report_type=report_type)
    report_filename = f"Reporte_Inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(report_filename)
    return report_filename

if __name__ == "__main__":
    db = DatabaseConnection()
    db.create_tables()  # Asegúrate de que las tablas se crean


class DatabaseConnection:
    def create_tables(self):
        sql_productos = '''
        CREATE TABLE IF NOT EXISTS productos (
            codigo TEXT PRIMARY KEY,
            descripcion TEXT,
            entradas_totales INTEGER DEFAULT 0,
            salidas_totales INTEGER DEFAULT 0,
            stock INTEGER DEFAULT 0,
            precio_compra REAL,
            precio_venta REAL,
            valor_total REAL
        )'''

        sql_entradas = '''
        CREATE TABLE IF NOT EXISTS entradas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT,
            descripcion TEXT,
            cantidad INTEGER,
            fecha TEXT,
            FOREIGN KEY (codigo) REFERENCES productos(codigo)
        )'''

        sql_salidas = '''
        CREATE TABLE IF NOT EXISTS salidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT,
            descripcion TEXT,
            cantidad INTEGER,
            fecha TEXT,
            FOREIGN KEY (codigo) REFERENCES productos(codigo)
        )'''

        with self.connect() as conn:
            cursor = conn.cursor()
            cursor.execute(sql_productos)
            cursor.execute(sql_entradas)
            cursor.execute(sql_salidas)
            conn.commit()

class ReporteView(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.db = DatabaseConnection()

    def init_ui(self):
        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        # Título y botones
        header_layout = QHBoxLayout()
        title_label = QLabel("Generación de Reportes")
        title_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        header_layout.addWidget(title_label)
        header_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))
        main_layout.addLayout(header_layout)

        # Filtros en línea horizontal
        filters_layout = QHBoxLayout()
        
        # Fecha inicio
        filters_layout.addWidget(QLabel("Fecha inicio:"))
        self.fecha_inicio = QDateEdit()
        self.fecha_inicio.setCalendarPopup(True)
        self.fecha_inicio.setDate(QDate.currentDate())
        filters_layout.addWidget(self.fecha_inicio)
        
        # Fecha fin
        filters_layout.addWidget(QLabel("Fecha fin:"))
        self.fecha_fin = QDateEdit()
        self.fecha_fin.setCalendarPopup(True)
        self.fecha_fin.setDate(QDate.currentDate())
        filters_layout.addWidget(self.fecha_fin)
        
        # Botón generar
        self.btn_generar_reporte = QPushButton("Generar Reporte Excel")
        self.btn_generar_reporte.clicked.connect(self.generar_reporte)
        filters_layout.addWidget(self.btn_generar_reporte)
        
        main_layout.addLayout(filters_layout)

        # Búsqueda de producto para historial
        search_group = QGroupBox("Historial de Productos")
        search_layout = QVBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar por código o descripción")
        self.search_input.textChanged.connect(self.buscar_historial)
        search_layout.addWidget(self.search_input)

        # Tabla de historial
        self.historial_table = QTableWidget()
        self.historial_table.setColumnCount(10)
        self.historial_table.setHorizontalHeaderLabels([
            "Fecha", "Código", "Descripción", "Tipo Mov.", 
            "Cantidad", "Stock", "Precio Compra", "Precio Venta",
            "Variación Precio", "Valor Total"
        ])
        search_layout.addWidget(self.historial_table)
        
        search_group.setLayout(search_layout)
        main_layout.addWidget(search_group)

    def buscar_historial(self, texto_busqueda):
        if len(texto_busqueda) >= 3:  # Búsqueda después de 3 caracteres
            with self.db.connect() as conn:
                cursor = conn.cursor()
                query = """
                SELECT p.codigo, p.descripcion, 
                        COALESCE(e.fecha, s.fecha) as fecha,
                        CASE
                            WHEN e.cantidad IS NOT NULL THEN 'Entrada'
                            WHEN s.cantidad IS NOT NULL THEN 'Salida'
                        END as tipo_movimiento,
                        COALESCE(e.cantidad, s.cantidad) as cantidad,
                        p.stock,
                        p.precio_compra,
                        p.precio_venta,
                        '' as variacion_precio,
                        COALESCE(e.cantidad, s.cantidad) * 
                        CASE
                            WHEN e.cantidad IS NOT NULL THEN p.precio_compra
                            WHEN s.cantidad IS NOT NULL THEN p.precio_venta
                        END as valor_total
                FROM productos p
                LEFT JOIN entradas e ON p.codigo = e.codigo
                LEFT JOIN salidas s ON p.codigo = s.codigo
                WHERE p.codigo LIKE ? OR p.descripcion LIKE ?
                ORDER BY fecha DESC
                LIMIT 100
                """
                cursor.execute(query, (f'%{texto_busqueda}%', f'%{texto_busqueda}%'))
                resultados = cursor.fetchall()

                self.historial_table.setRowCount(len(resultados))
                for row, dato in enumerate(resultados):
                    for col, value in enumerate(dato):
                        item = QTableWidgetItem(str(value))
                        self.historial_table.setItem(row, col, item)

    def generar_reporte(self):
        try:
            fecha_inicio = self.fecha_inicio.date().toPyDate()
            fecha_fin = self.fecha_fin.date().toPyDate()
            
            # Obtener datos de la base de datos
            with self.db.connect() as conn:
                cursor = conn.cursor()
                
                # Consulta para inventario con cálculos
                cursor.execute("""
                    SELECT 
                        p.codigo,
                        p.descripcion,
                        COALESCE(SUM(e.cantidad), 0) as entradas,
                        COALESCE(SUM(s.cantidad), 0) as salidas,
                        p.stock,
                        p.precio_compra,
                        p.precio_venta,
                        COALESCE(SUM(e.cantidad), 0) * p.precio_compra as valor_compra_total,
                        COALESCE(SUM(s.cantidad), 0) * p.precio_venta as valor_venta_total,
                        (COALESCE(SUM(s.cantidad), 0) * p.precio_venta) - 
                        (COALESCE(SUM(e.cantidad), 0) * p.precio_compra) as utilidad
                    FROM productos p
                    LEFT JOIN entradas e ON p.codigo = e.codigo
                    LEFT JOIN salidas s ON p.codigo = s.codigo
                    GROUP BY p.codigo, p.descripcion, p.stock, p.precio_compra, p.precio_venta
                """)
                data_inventario = [list(row) for row in cursor.fetchall()]
                
                # Consulta para entradas
                cursor.execute("""
                    SELECT 
                        e.fecha,
                        e.codigo,
                        p.descripcion,
                        e.cantidad,
                        p.precio_compra,
                        e.cantidad * p.precio_compra as valor_total
                    FROM entradas e
                    JOIN productos p ON e.codigo = p.codigo
                    WHERE e.fecha BETWEEN ? AND ?
                    ORDER BY e.fecha
                """, (fecha_inicio, fecha_fin))
                data_entradas = [list(row) for row in cursor.fetchall()]
                
                # Consulta para salidas
                cursor.execute("""
                    SELECT 
                        s.fecha,
                        s.codigo,
                        p.descripcion,
                        s.cantidad,
                        p.precio_venta,
                        s.cantidad * p.precio_venta as valor_total
                    FROM salidas s
                    JOIN productos p ON s.codigo = p.codigo
                    WHERE s.fecha BETWEEN ? AND ?
                    ORDER BY s.fecha
                """, (fecha_inicio, fecha_fin))
                data_salidas = [list(row) for row in cursor.fetchall()]
            
            # Verificar que los datos no estén vacíos
            if not data_inventario:
                data_inventario = []
            if not data_entradas:
                data_entradas = []
            if not data_salidas:
                data_salidas = []

            # Generar el reporte
            generator = ReporteExcelGenerator()
            wb = generator.generate_report(
                data_inventario=data_inventario,
                data_entradas=data_entradas,
                data_salidas=data_salidas,
                start_date=fecha_inicio,
                end_date=fecha_fin
            )
            
            # Guardar el archivo
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Reporte Excel",
                "",
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                wb.save(file_path)
                print(f"Reporte guardado en: {file_path}")
                
        except Exception as e:
            print(f"Error al generar el reporte: {str(e)}")
            # Aquí podrías mostrar un mensaje de error al usuario usando QMessageBox

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def generar_reporte_excel(fecha_inicio, fecha_fin, db):
    wb = Workbook()
    
    # Estilos
    font_bold = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    

    # Hoja de Productos
    ws_productos = wb.active
    ws_productos.title = "Productos"
    
    # Encabezado

    ws_productos.merge_cells('B1:F1')
    ws_productos['B1'] = "MULTIMUEBLES LA PLATA - HUILA"
    ws_productos['B1'].font = Font(size=14, bold=True)
    ws_productos['B1'].alignment = alignment_center
    
    ws_productos.merge_cells('B2:F2')
    ws_productos['B2'] = "INVENTARIO DE MATERIA PRIMA"
    ws_productos['B2'].font = Font(size=12, bold=True)
    ws_productos['B2'].alignment = alignment_center
    
    ws_productos.merge_cells('B3:F3')
    ws_productos['B3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_productos['B3'].alignment = alignment_center
    
    ws_productos.merge_cells('B4:F4')
    ws_productos['B4'] = f"Periodo del reporte: {fecha_inicio} a {fecha_fin}"
    ws_productos['B4'].alignment = alignment_center
    
    # Encabezados de la tabla
    headers = ["Código", "Descripción", "Entradas", "Salidas", "Stock Actual", "Precio Compra", "Precio Venta"]
    ws_productos.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_productos.cell(row=6, column=col_num)
        cell.value = header
        cell.font = font_bold
        cell.alignment = alignment_center
        cell.fill = fill_header
        cell.border = border_thin
        ws_productos.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Obtener datos de la base de datos
    with db.connect() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.codigo, p.descripcion, 
                    SUM(CASE WHEN e.fecha BETWEEN ? AND ? THEN e.cantidad ELSE 0 END) as entradas,
                    SUM(CASE WHEN s.fecha BETWEEN ? AND ? THEN s.cantidad ELSE 0 END) as salidas,
                    p.stock, p.precio_compra, p.precio_venta
            FROM productos p
            LEFT JOIN entradas e ON p.codigo = e.codigo
            LEFT JOIN salidas s ON p.codigo = s.codigo
            GROUP BY p.codigo, p.descripcion
        """, (fecha_inicio, fecha_fin, fecha_inicio, fecha_fin))
        datos = cursor.fetchall()
    
    # Añadir datos a la tabla
    for row in datos:
        ws_productos.append(row)
    
    # Ajustar ancho de columnas
    for col in ws_productos.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_productos.column_dimensions[column].width = adjusted_width
    
    # Crear hoja de Entradas
    ws_entradas = wb.create_sheet(title="Entradas")
    
    # Encabezado
    
    ws_entradas.merge_cells('B1:F1')
    ws_entradas['B1'] = "MULTIMUEBLES LA PLATA - HUILA"
    ws_entradas['B1'].font = Font(size=14, bold=True)
    ws_entradas['B1'].alignment = alignment_center
    
    ws_entradas.merge_cells('B2:F2')
    ws_entradas['B2'] = "INVENTARIO DE MATERIA PRIMA"
    ws_entradas['B2'].font = Font(size=12, bold=True)
    ws_entradas['B2'].alignment = alignment_center
    
    ws_entradas.merge_cells('B3:F3')
    ws_entradas['B3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_entradas['B3'].alignment = alignment_center
    
    ws_entradas.merge_cells('B4:F4')
    ws_entradas['B4'] = f"Periodo del reporte: {fecha_inicio} a {fecha_fin}"
    ws_entradas['B4'].alignment = alignment_center
    
    # Encabezados de la tabla
    headers_entradas = ["Fecha", "Código", "Descripción", "Cantidad", "Precio Compra", "Total"]
    ws_entradas.append(headers_entradas)
    
    for col_num, header in enumerate(headers_entradas, 1):
        cell = ws_entradas.cell(row=6, column=col_num)
        cell.value = header
        cell.font = font_bold
        cell.alignment = alignment_center
        cell.fill = fill_header
        cell.border = border_thin
        ws_entradas.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Obtener datos de la base de datos
    with db.connect() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT e.fecha, e.codigo, p.descripcion, e.cantidad, p.precio_compra, (e.cantidad * p.precio_compra) as total
            FROM entradas e
            JOIN productos p ON e.codigo = p.codigo
            WHERE e.fecha BETWEEN ? AND ?
        """, (fecha_inicio, fecha_fin))
        datos_entradas = cursor.fetchall()
    
    # Añadir datos a la tabla
    for row in datos_entradas:
        ws_entradas.append(row)
    
    # Ajustar ancho de columnas
    for col in ws_entradas.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_entradas.column_dimensions[column].width = adjusted_width
    
    # Crear hoja de Salidas
    ws_salidas = wb.create_sheet(title="Salidas")
    
    # Encabezado
    ws_salidas.merge_cells('B1:F1')
    ws_salidas['B1'] = "MULTIMUEBLES LA PLATA - HUILA"
    ws_salidas['B1'].font = Font(size=14, bold=True)
    ws_salidas['B1'].alignment = alignment_center
    
    ws_salidas.merge_cells('B2:F2')
    ws_salidas['B2'] = "INVENTARIO DE MATERIA PRIMA"
    ws_salidas['B2'].font = Font(size=12, bold=True)
    ws_salidas['B2'].alignment = alignment_center
    
    ws_salidas.merge_cells('B3:F3')
    ws_salidas['B3'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_salidas['B3'].alignment = alignment_center
    
    ws_salidas.merge_cells('B4:F4')
    ws_salidas['B4'] = f"Periodo del reporte: {fecha_inicio} a {fecha_fin}"
    ws_salidas['B4'].alignment = alignment_center
    
    # Encabezados de la tabla
    headers_salidas = ["Fecha", "Código", "Descripción", "Cantidad", "Precio Venta", "Total"]
    ws_salidas.append(headers_salidas)
    
    for col_num, header in enumerate(headers_salidas, 1):
        cell = ws_salidas.cell(row=6, column=col_num)
        cell.value = header
        cell.font = font_bold
        cell.alignment = alignment_center
        cell.fill = fill_header
        cell.border = border_thin
        ws_salidas.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Obtener datos de la base de datos
    with db.connect() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT s.fecha, s.codigo, p.descripcion, s.cantidad, p.precio_venta, (s.cantidad * p.precio_venta) as total
            FROM salidas s
            JOIN productos p ON s.codigo = p.codigo
            WHERE s.fecha BETWEEN ? AND ?
        """, (fecha_inicio, fecha_fin))
        datos_salidas = cursor.fetchall()
    
    # Añadir datos a la tabla
    for row in datos_salidas:
        ws_salidas.append(row)
    
    # Ajustar ancho de columnas
    for col in ws_salidas.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_salidas.column_dimensions[column].width = adjusted_width
    
    return wb
